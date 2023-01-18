import sys
import datetime
import openpyxl
import xlrd,xlwt
from Database_project.project.data_base_ import db
from Database_project.project.data_base_.Models import Tarifs,Client,Expert
from Database_project.project.data_base_.client_data  import regex1
from flask import Flask,render_template,url_for,flash,redirect,request,Blueprint,send_from_directory
import os

def ex(name):
    if name == None:
        return 0
    if name == '':
        return 0
    cli=Expert.query.filter_by(full=str(name.lower())).first()
    if cli is not None:
        return cli.id
    else:
        return 0

def good2(av):
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Bien')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    #wb.save('C:/Users/user/Downloads/Telegram Desktop/Tarifbien.xls')

def failed(av):
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Anomalie')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    filename='Tarif_failed.xls'
    file_path = os.path.join(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'),filename)
    loc=str(file_path)
    # set the file path
    #uploaded_file.save(file_path)
    wb.save(loc)
    return send_from_directory(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'), filename,as_attachment=True)

def arif(loc):
    #try:
        anomalie=[]
        Bien=[]
        a=[]
        b=[]
        p=0
        wb_obj = openpyxl.load_workbook(loc,data_only=True)
        sheet=wb_obj.active
        
        '''if sheet["W"][0].value!='CODE POSTAL'and sheet["AB"][0].value!='Tel principal client'and sheet["J"][0].value!='Ref code client- service comptabilité'and sheet["P"][0].value!='Fonction responsable':
                    return False'''
        for i in range(0,sheet.max_row):
            #print(loc)
            if sheet["J"][i].value!='PROSPECT':
                    v=sheet["M"][i].value
                    if type(v)==int: 
                        if v  not in a:
                            cli=Client.query.filter_by(reference=int(sheet["M"][i].value)).first()
                            a.append(v)
                            if cli:
                                cli2=Tarifs.query.filter_by(reference_client=cli.id).first()
                            if cli and  cli2 == None :

                                tarif=Tarifs()
                                edl_prix_std=regex1(sheet["BN"][i].value,'dec')   
                                if edl_prix_std=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_prix_std  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_prix_std=edl_prix_std
                                edl_appt_prix_f1=regex1(sheet["BO"][i].value,'dec')
                                if edl_appt_prix_f1=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_appt_prix_f1  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_appt_prix_f1=edl_appt_prix_f1
                                edl_appt_prix_f2=regex1(sheet["BP"][i].value,'dec')
                                if edl_appt_prix_f2=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_appt_prix_f2  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_appt_prix_f2=edl_appt_prix_f2
                                edl_appt_prix_f3=regex1(sheet["BQ"][i].value,'dec')
                                if edl_appt_prix_f3=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_appt_prix_f3  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_appt_prix_f3=edl_appt_prix_f3
                                edl_appt_prix_f4=regex1(sheet["BR"][i].value,'dec')
                                if edl_appt_prix_f4=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_appt_prix_f4  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_appt_prix_f4=edl_appt_prix_f4
                                edl_appt_prix_f5=regex1(sheet["BS"][i].value,'dec')
                                if edl_appt_prix_f5=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_appt_prix_f5  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_appt_prix_f5=edl_appt_prix_f5
                                edl_appt_prix_f6=regex1(sheet["BT"][i].value,'dec')
                                if edl_appt_prix_f6=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_appt_prix_f6  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_appt_prix_f6=edl_appt_prix_f6
                                edl_pav_villa_prix_t1=regex1(sheet["BU"][i].value,'dec')
                                if edl_pav_villa_prix_t1=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_pav_villa_prix_t1  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_pav_villa_prix_t1=edl_pav_villa_prix_t1
                                edl_pav_villa_prix_t2=regex1(sheet["BV"][i].value,'dec')
                                if edl_pav_villa_prix_t2=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_pav_villa_prix_t2  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_pav_villa_prix_t2=edl_pav_villa_prix_t2
                                edl_pav_villa_prix_t3=regex1(sheet["BW"][i].value,'dec')
                                if edl_pav_villa_prix_t3=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_pav_villa_prix_t3  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_pav_villa_prix_t3=edl_pav_villa_prix_t3
                                edl_pav_villa_prix_t4=regex1(sheet["BX"][i].value,'dec')
                                if edl_pav_villa_prix_t4=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_pav_villa_prix_t4  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_pav_villa_prix_t4=edl_pav_villa_prix_t4
                                edl_pav_villa_prix_t5=regex1(sheet["BY"][i].value,'dec')
                                if edl_pav_villa_prix_t5=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_pav_villa_prix_t5  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_pav_villa_prix_t5=edl_pav_villa_prix_t5
                                edl_pav_villa_prix_t6=regex1(sheet["BZ"][i].value,'dec')
                                if edl_pav_villa_prix_t6=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_pav_villa_prix_t6  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_pav_villa_prix_t6=edl_pav_villa_prix_t6
                                edl_pav_villa_prix_t7=regex1(sheet["CA"][i].value,'dec')
                                if edl_pav_villa_prix_t7=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_pav_villa_prix_t7  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_pav_villa_prix_t7=edl_pav_villa_prix_t7
                                edl_pav_villa_prix_t8=regex1(sheet["CB"][i].value,'dec')
                                if edl_pav_villa_prix_t8=="Failed":
                                    reason='erreur  de numero dans la colonne  edl_pav_villa_prix_t8  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.edl_pav_villa_prix_t8=edl_pav_villa_prix_t8
                                chif_appt_prix_stu=regex1(sheet["CC"][i].value,'dec')
                                if chif_appt_prix_stu=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_appt_prix_stu  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_appt_prix_stu=chif_appt_prix_stu
                                chif_appt_prix_f1=regex1(sheet["CD"][i].value,'dec')
                                if chif_appt_prix_f1=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_appt_prix_f1  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_appt_prix_f1=chif_appt_prix_f1
                                chif_appt_prix_f2=regex1(sheet["CE"][i].value,'dec')
                                if chif_appt_prix_f2=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_appt_prix_f2  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_appt_prix_f2=chif_appt_prix_f2
                                chif_appt_prix_f3=regex1(sheet["CF"][i].value,'dec')
                                if chif_appt_prix_f3=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_appt_prix_f3  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_appt_prix_f3=chif_appt_prix_f3
                                chif_appt_prix_f4=regex1(sheet["CG"][i].value,'dec')
                                if chif_appt_prix_f4=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_appt_prix_f4  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_appt_prix_f4=chif_appt_prix_f4
                                chif_appt_prix_f5=regex1(sheet["CH"][i].value,'dec')
                                if chif_appt_prix_f5=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_appt_prix_f5  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_appt_prix_f5=chif_appt_prix_f5
                                #tarif.chif_appt_prix_f6=regex1(sheet["CI"][i].value,'dec')
                                chif_pav_villa_prix_t1=regex1(sheet["CI"][i].value,'dec')
                                if chif_pav_villa_prix_t1=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_pav_villa_prix_t1  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_pav_villa_prix_t1=chif_pav_villa_prix_t1
                                chif_pav_villa_prix_t2=regex1(sheet["CJ"][i].value,'dec')
                                if chif_pav_villa_prix_t2=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_pav_villa_prix_t2  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_pav_villa_prix_t2=chif_pav_villa_prix_t2
                                chif_pav_villa_prix_t3=regex1(sheet["CK"][i].value,'dec')
                                if chif_pav_villa_prix_t3=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_pav_villa_prix_t3  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_pav_villa_prix_t3=chif_pav_villa_prix_t3
                                chif_pav_villa_prix_t4=regex1(sheet["CL"][i].value,'dec')
                                if chif_pav_villa_prix_t4=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_pav_villa_prix_t4  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_pav_villa_prix_t4=chif_pav_villa_prix_t4
                                chif_pav_villa_prix_t5=regex1(sheet["CM"][i].value,'dec')
                                if chif_pav_villa_prix_t5=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_pav_villa_prix_t5  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_pav_villa_prix_t5=chif_pav_villa_prix_t5
                                chif_pav_villa_prix_t6=regex1(sheet["CN"][i].value,'dec')
                                if chif_pav_villa_prix_t6=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_pav_villa_prix_t6  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_pav_villa_prix_t6=chif_pav_villa_prix_t6
                                chif_pav_villa_prix_t7=regex1(sheet["CO"][i].value,'dec')
                                if chif_pav_villa_prix_t7=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_pav_villa_prix_t7  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_pav_villa_prix_t7=chif_pav_villa_prix_t7
                                chif_pav_villa_prix_t8=regex1(sheet["CP"][i].value,'dec')
                                if chif_pav_villa_prix_t8=="Failed":
                                    reason='erreur  de numero dans la colonne  chif_pav_villa_prix_t8  ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                    sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                    sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                    sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                    sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                    sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                    sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                    sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                    sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                    sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                    sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                    sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                    sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                    sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                    sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                    sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                    sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                    sheet["CQ"][i].value,reason])
                                    continue
                                else:
                                    tarif.chif_pav_villa_prix_t8=chif_pav_villa_prix_t8
                                #tarif.prix_autre=

                                #tarif.code_tva=regex1(sheet["CP"][i].value,'str1')

                                tarif.taux_meuble=regex1(sheet["BM"][i].value,'perc')

                                tarif.referent_as_client=ex(sheet["AU"][i].value)

                                tarif.com_as_sur_ca_client=regex1(sheet["AV"][i].value,'perc')

                                tarif.cell_dev_ref_responsable=ex(sheet["AW"][i].value)

                                tarif.com_cell_dev_ref_responsable=regex1(sheet["AX"][i].value,'perc')

                                tarif.cell_dev_ref_agent=ex(sheet["AY"][i].value)

                                tarif.com_cell_dev_ref_agent =regex1(sheet["AZ"][i].value,'perc')

                                tarif.cell_tech_ref_agent =ex(sheet["BA"][i].value)

                                tarif.com_cell_tech_Ref_agent =regex1(sheet["BB"][i].value,'perc')

                                tarif.cell_tech_ref_responsable =ex(sheet["BC"][i].value)

                                tarif.com_cell_tech_ref_responsable=regex1(sheet["BD"][i].value,'perc')

                                tarif.cell_tech_ref_suiveur=ex(sheet["BE"][i].value)

                                tarif.com_cell_tech_ref_suiveur=regex1(sheet["BF"][i].value,'perc')

                                tarif.cell_planif_ref_responsable=ex(sheet["BG"][i].value)

                                tarif.com_cell_planif_ref_responsable=regex1(sheet["BH"][i].value,'perc')

                                tarif.cell_planif_ref_suiveur=ex(sheet["BI"][i].value)

                                tarif.com_cell_planif_ref_suiveur=regex1(sheet["BJ"][i].value,'perc')

                                tarif.cell_planif_ref_agent_saisie=ex(sheet["BK"][i].value)

                                tarif.com_cell_planif_ref_agent_saisie=regex1(sheet["BL"][i].value,'perc')

                                tarif.commentaire_libre =regex1(sheet["CQ"][i].value,'str1')
                                tarif.reference_client=cli.id
                                db.session.add(tarif)
                                db.session.commit()
                                
                                Bien.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                                sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                                sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                                sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                                sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                                sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                                sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                                sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                                sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                                sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                                sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                                sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                                sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                                sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                                sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                                sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                                sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                                sheet["CQ"][i].value])
                    else:     
                        reason='erreur  de numero dans la Reference'
                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                        sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                        sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                        sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                        ,sheet["AB"][i].value,sheet["AC"][i].value,sheet["AD"][i].value,sheet["AE"][i].value,
                        sheet["AF"][i].value,sheet["AG"][i].value,sheet["AH"][i].value,
                        sheet["AI"][i].value,sheet["AJ"][i].value,sheet["AK"][i].value,sheet["AL"][i].value,
                        sheet["AM"][i].value,sheet["AN"][i].value,sheet["AO"][i].value,sheet["AP"][i].value,
                        sheet["AQ"][i].value,sheet["AR"][i].value,sheet["AS"][i].value,sheet["AT"][i].value,
                        sheet["AU"][i].value,sheet["AV"][i].value,sheet["AW"][i].value,sheet["AX"][i].value,
                        sheet["AY"][i].value,sheet["AZ"][i].value,sheet["BA"][i].value,sheet["BB"][i].value,
                        sheet["BC"][i].value,sheet["BD"][i].value,sheet["BE"][i].value,sheet["BF"][i].value,
                        sheet["BG"][i].value,sheet["BH"][i].value,sheet["BI"][i].value,sheet["BJ"][i].value,
                        sheet["BK"][i].value,sheet["BL"][i].value,sheet["BM"][i].value,sheet["BN"][i].value,
                        sheet["BO"][i].value,sheet["BP"][i].value,sheet["BQ"][i].value,sheet["BR"][i].value,
                        sheet["BS"][i].value,sheet["BT"][i].value,sheet["BU"][i].value,sheet["BV"][i].value,
                        sheet["BW"][i].value,sheet["BX"][i].value,sheet["BY"][i].value,sheet["BZ"][i].value,
                        sheet["CA"][i].value,sheet["CB"][i].value,sheet["CC"][i].value,sheet["CD"][i].value,
                        sheet["CE"][i].value,sheet["CF"][i].value,sheet["CG"][i].value,sheet["CH"][i].value,
                        sheet["CI"][i].value,sheet["CJ"][i].value,sheet["CK"][i].value,sheet["CL"][i].value,
                        sheet["CM"][i].value,sheet["CN"][i].value,sheet["CO"][i].value,sheet["CP"][i].value,
                        sheet["CQ"][i].value,reason])      
        if anomalie == []:
            return True
        else:
            return failed(anomalie)
        '''if Bien == []:
            print(Bien)
        else:
            good2(Bien)'''
    #except:
     #   return  False
