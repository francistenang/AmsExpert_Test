from project.data_base_ import db
from project.data_base_.Models import Tarifs,Mission,Client,Expert,Client_History,prospect,prospect_History,Expert_History,Tarif_base
import xlrd
import openpyxl

def insert_client(A,loc):
    

    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)
    
    sheet.cell_value(0, 0)
    for i in range(0,400):
        name=sheet.row_values(i+1)
        
        if A == 'Bailleur':
            
            if name[3] == '' or name[3] == 'XX':
                print('no data here')
            else:
                cli=Client.query.filter_by(nom=str(name[3].lower())).first()#arrange sort by reference
                if cli is None:
                    client=Client(TYPE=A,societe=name[1].lower(),sexe=name[2].lower(),nom=name[3].lower(),email='test@gmail.com',numero='6777650822',siret='22222222222')
                    db.session.add(client)
                    db.session.commit()   
                    client.reference=str(name[0])
                    db.session.commit()
                    history=Client_History(client_id=client.id,adresse=name[4],cp=name[6],ville=name[7],pays='Britain')
                    db.session.add(history)
                    db.session.commit()
                else:
                    print('already exist')

        if A == 'Prop':
            if name[43] == '' or name[43] == 'XX':
                print('no data here')
            else:
                cli=Client.query.filter_by(NOM=str(name[43].lower())).first()
                if cli is None:
                    cli2=Client('',A,'',name[42].lower(),name[43].lower(),'','','','','','','','','')
                    db.session.add(cli2)
                    db.session.commit()
                else:
                    print('already exist')

        if A == 'Locataire':
            if name[20] == '' or name[20] == 'XX':
                print('no data here')
            else:
                pros1=prospect.query.filter_by(reference=str(name[18])).first()    
                if pros1 is None:
                    client=prospect(A,'Amexpert',name[19].lower(),name[20].lower(),'test@gmail.com','675337250')
                    db.session.add(client)
                    db.session.commit()
                    client.reference=str(name[18])
                    db.session.commit()
                    user_his=prospect_History(prospect=client.id,adresse=name[21],cp=name[23],ville=name[24],pays='Britain')
                    db.session.add(user_his)
                    db.session.commit()

                else:
                    print('already exist')


def expert__(A,loc):

    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)

    sheet.cell_value(0, 0)
    for i in range(0,400):
        name=sheet.row_values(i+1)
        if A == 'Interv':
            if name[17] == '' or name[17] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(nom=str(name[17].lower())).first()
                if cli is None:
                    expert=Expert(sexe=name[16],nom=name[17],numero='222000000',TYPE=A, email='test@gmail.com')    
                    db.session.add(expert)
                    db.session.commit()
                    history=Expert_History(expert_id=expert.id)
                    db.session.add(history)
                    db.session.commit()
                else:
                    print('already exist')

        if A == 'CONCESS':
            if name[11] == '' or name[11] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(nom=str(name[11].lower())).first()
                if cli is None:
                    expert=Expert(sexe=name[10],nom=name[11],numero='222000000',TYPE=A, email='test@gmail.com')      
                    db.session.add(expert)
                    db.session.commit()
                    history=Expert_History(expert_id=expert.id)
                    db.session.add(history)
                    db.session.commit()
                else:
                    print('already exist')

        if A == 'Manager_chiffrage':
            if name[37] == '' or name[37] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(NOM=str(name[37].lower())).first()
                if cli is None:
                    expert=Expert('',name[37].lower(),A,'','')
                    db.session.add(expert)
                    db.session.commit()
                else:
                    print('already exist')
        
        if A == 'Agent_chiffrage':
            if name[39] == '' or name[39] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(NOM=str(name[39].lower())).first()
                if cli is None:
                    expert=Expert('',name[39].lower(),A,'','')
                    db.session.add(expert)
                    db.session.commit()
                else:
                    print('already exist')

        if A == 'Respon Cell Dev':
            if name[69] == '' or name[69] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(NOM=str(name[69].lower()).lower()).first()
                if cli is None:
                    expert=Expert('',name[69].lower(),A,'','')
                    db.session.add(expert)
                    db.session.commit()
                else:
                    print('already exist')

        

        if A == 'agent Cell Dev':
            if name[71] == '' or name[71] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(NOM=str(name[71].lower())).first()
                if cli is None:
                    expert=Expert('',name[71].lower(),A,'','')
                    db.session.add(expert)
                    db.session.commit()
                else:
                    print('already exist')

        if A == 'Agent CellTech':
            if name[73] == '' or name[73] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(NOM=str(name[73].lower())).first()
                if cli is None:
                    expert=Expert('',name[73].lower(),A,'','')
                    db.session.add(expert)
                    db.session.commit()
                else:
                    print('already exist')

        if A == 'Respon Cell Tech':
            if name[75] == '' or name[75] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(NOM=str(name[75].lower())).first()
                if cli is None:
                    expert=Expert('',name[75].lower(),A,'','')
                    db.session.add(expert)
                    db.session.commit()
                else:
                    print('already exist')

        if A == 'Suiveur Cell Tech':
            if name[77] == '' or name[77] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(NOM=str(name[77].lower())).first()
                if cli is None:
                    expert=Expert('',name[77].lower(),A,'','')
                    db.session.add(expert)
                    db.session.commit()
                else:
                    print('already exist')

        if A == 'Respon Cell Planif':
            if name[79] == '' or name[79] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(NOM=str(name[79].lower())).first()
                if cli is None:
                    expert=Expert('',name[79].lower(),A,'','')
                    db.session.add(expert)
                    db.session.commit()
                else:
                    print('already exist')

        if A == 'Suiveur Cell Planif':
            if name[81] == '' or name[81] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(NOM=str(name[81].lower())).first()
                if cli is None:
                    expert=Expert('',name[81].lower(),A,'','')
                    db.session.add(expert)
                    db.session.commit()
                else:
                    print('already exist')

        if A == 'Agent saisie Cell Planif':
            if name[83] == '' or name[83] == 'XX':
                print('no data here')
            else:
                cli=Expert.query.filter_by(NOM=str(name[83].lower())).first()
                if cli is None:
                    expert=Expert('',name[83].lower(),A,'','')
                    db.session.add(expert)
                    db.session.commit()
                else:
                    print('already exist')
        

def tarif_client(loc):
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)

    sheet.cell_value(0, 0)
    for i in range(0,10):
        name=sheet.row_values(i+1)
        tarif =Tarifs.query.filter_by(reference_client=int(name[1])).first()
        if tarif is None:
            taf_base =Tarifs(reference_client=int(name[1]),code_tva=int(name[2]),
            referent_as_client=int(name[3]),com_as_sur_ca_client=name[4],cell_dev_ref_responsable=name[5],
            com_cell_dev_ref_responsable=name[6],Cell_dev_ref_agent=int(name[7]),com_cell_dev_ref_agent=name[8],
            Cell_tech_Ref_agent=int(name[9]),com_cell_tech_Ref_agent=name[10],CELL_TECH_ref_responsable=int(name[11]),
            COM_CELL_TECH_ref_responsable=name[12],CELL_TECH_ref_suiveur=int(name[13]),
            com_CELL_TECH_ref_suiveur=name[14],CELL_planif_ref_responsable=int(name[15]),
            com_CELL_planif_ref_responsable=name[16],CELL_PLANIF_ref_suiveur=int(name[17]),
            com_CELL_PLANIF_ref_suiveur=name[18],CELL_PLANIF_ref_agent_saisie=int(name[19]),
            com_CELL_PLANIF_ref_agent_saisie=name[20],taux_meuble=name[21],EDL_PRIX_STD=name[22],
            EDL_APPT_prix_F1=name[23],EDL_APPT_prix_F2=name[24],EDL_APPT_prix_F3=name[25],EDL_APPT_prix_F4=name[26],
            EDL_APPT_prix_F5=name[27],EDL_APPT_prix_F6=name[28],EDL_PAV_villa_prix_T1=name[29], EDL_PAV_villa_prix_T2=name[30],
            EDL_PAV_villa_prix_T3=name[31],EDL_PAV_villa_prix_T4=name[32],EDL_PAV_villa_prix_T5=name[33],EDL_PAV_villa_prix_T6=name[34],
            EDL_PAV_villa_prix_T7=name[35],EDL_PAV_villa_prix_T8=name[36],CHIF_APPT_prix_stu=name[37],
            CHIF_APPT_prix_F1=name[38],CHIF_APPT_prix_F2=name[39],CHIF_APPT_prix_F3=name[40],
            CHIF_APPT_prix_F4=name[41],CHIF_APPT_prix_F5=name[42],CHIF_PAV_villa_prix_T1=name[43],
            CHIF_PAV_villa_prix_T2=name[44],CHIF_PAV_villa_prix_T3=name[45],CHIF_PAV_villa_prix_T4=name[46],
            CHIF_PAV_villa_prix_T5=name[47],CHIF_PAV_villa_prix_T6=name[48],CHIF_PAV_villa_prix_T7=name[49],
            CHIF_PAV_villa_prix_T8=name[50])
            
def Missions(loc):

    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)

    sheet.cell_value(0, 0)
    for i in range(0,50):
        name=sheet.row_values(i+1)
        
        client1=Client.query.filter_by(nom=str(name[3].lower())).first()
        if client1 is not None:
            name[3]=client1.id
        if name[3] =='XX' :
            name[3]=0
        if name[3] ==''  :
            name[3]=0

        client2=Client.query.filter_by(nom=str(name[43].lower())).first()
        if client2 is not None:
            name[43]=client2.id
        if name[43] =='XX' :
            name[43]=0
        if name[43] ==''  :
            name[43]=0

        client3=Client.query.filter_by(nom=str(name[20].lower())).first() 
        if client3 is not None:
            name[20]=client3.id
        if name[20] =='XX' :
            name[20]=0
        if name[20] ==''  :
            name[20]=0

        expert1=Expert.query.filter_by(nom=str(name[17].lower())).first()
        if expert1 is not None:
            name[17]=expert1.id
        if name[17] =='XX' :
            name[17]=0
        if name[17] ==''  :
            name[17]=0

        expert2=Expert.query.filter_by(nom=str(name[11].lower())).first()
        if expert2 is not None:
            name[11]=expert2.id
        if name[11] =='XX' :
            name[11]=0
        if name[11] ==''  :
            name[11]=0

        expert3=Expert.query.filter_by(nom=str(name[37].lower())).first()
        if expert3 is not None:
            name[37]=expert3.id
        if name[37] =='XX'  :
            name[37]=0
        if name[37] ==''  :
            name[37]=0

        expert4=Expert.query.filter_by(nom=str(name[39].lower())).first()
        if expert4 is not None:
            name[39]=expert4.id
        if name[39] =='XX' :
            name[39]=0
        if name[39] ==''  :
            name[39]=0

        expert5=Expert.query.filter_by(nom=str(name[69].lower())).first()
        if expert5 is not None:
            name[69]=expert5.id
        if name[69] =='XX'  :
            name[69]=0
        if name[69] ==''  :
            name[69]=0

        expert6=Expert.query.filter_by(nom=str(name[71].lower())).first()
        if expert6 is not None:
            name[71]=expert6.id
        if name[71] =='XX'   :
            name[71]=0
        if name[71] ==''  :
            name[71]=0

        expert7=Expert.query.filter_by(nom=str(name[73].lower())).first()
        if expert7 is not None:
            name[73]=expert7.id
        if name[73] =='XX' :
            name[73]=0
        if name[73] ==''  :
            name[73]=0

        expert8=Expert.query.filter_by(nom=str(name[75].lower())).first()
        if expert8 is not None:
            name[75]=expert8.id
        if name[75] =='XX'  :
            name[75]=0
        if name[75] ==''  :
            name[75]=0

        expert9=Expert.query.filter_by(nom=str(name[77].lower())).first()
        if expert9 is not None:
            name[77]=expert9.id
        if name[77] =='XX' :
            name[77]=0
        if name[77] ==''  :
            name[77]=0

        expert10=Expert.query.filter_by(nom=str(name[79].lower())).first()
        if expert10 is not None:
            name[79]=expert10.id
        if name[79] =='XX'   :
            name[79]=0
        if name[79] ==''  :
            name[79]=0

        expert11=Expert.query.filter_by(nom=str(name[81].lower())).first()
        if expert11 is not None:
            name[81]=expert11.id
        if name[81] =='XX' :
            name[81]=0
        if name[81] ==''  :
            name[81]=0

        expert12=Expert.query.filter_by(nom=str(name[83].lower())).first()
        if expert12 is not None:
            name[83]=expert12.id
        if name[83] =='XX':
            name[83]=0
        if name[83] ==''  :
            name[83]=0
        
        print('ok')
        mission=Mission(name[3],name[8],name[11],name[12],name[13],name[14],name[15],name[17],name[20],name[25],name[26],name[27],name[28],name[29],name[30],name[31],name[32],name[33],name[34],name[35],name[36],name[37],name[38],name[39],
        name[40],name[41],name[43],name[44],name[45],name[46],name[47],name[48],name[49],name[50],name[51],name[52],name[53],name[54],name[55],name[56],name[57],name[58],name[59],name[60],
        name[61],name[62],name[63],name[64],name[65],name[66],name[67],name[68],name[69],name[70],name[71],name[72],name[73],name[74],name[75],name[76],name[77],name[78],name[79],name[80],name[81],
        name[82],name[83],name[84])

        db.session.add(mission)
        db.session.commit()

        
def mission_date(loc):
    wb_obj = openpyxl.load_workbook(loc)
    sheet=wb_obj.active
    mission_=list(Mission.query.all())
    for i in range(1,40):
        for a in range(0,40):
            mission_[a].DATE_REALISE_EDL =sheet["M"][i].value
            mission_[a].Date_chiffrage_regle =sheet["AG"][i].value
            mission_[a].DATE_FACTURE = sheet["AP"][i].value
            mission_[a].DATE_FACT_REGLEE = sheet["AS"][i].value
            db.session.commit()





#secondddddddddddddddddddd


def insert_client(A,loc):
    
    wb_obj = openpyxl.load_workbook(loc)
    sheet=wb_obj.active

    

    for i in range(1,sheet.max_row):
        
        if A == 'professionnel':
            
            if sheet["D"][i].value == '' or sheet["D"][i].value == 'XX':
                print('no data here')
            else:
                
                cli=Client.query.filter_by(reference=int(sheet["A"][i].value)).first()
                if cli is None:
                    client=Client(reference=int(sheet["A"][i].value),TYPE=A,societe=sheet["B"][i].value,titre=sheet["C"][i].value,nom=sheet["D"][i].value.lower(),date_creation=sheet["K"][i].value)
                    db.session.add(client)
                    db.session.commit()   
                    history=Client_History(client_id=client.id,adresse1=sheet["E"][i].value,adresse2=sheet["F"][i].value,cp=sheet["G"][i].value,ville=sheet["H"][i].value,login_extranet=sheet["M"][i].value,mpd_extranet=sheet["N"][i].value)
                    db.session.add(history)
                    db.session.commit()
                    if sheet["L"][i].value != None :
                        history.date=sheet["L"][i].value
                        db.session.commit()

                else:
                    return 'This data already exist'

        
#to be fixed

def expert__(loc):

    wb_obj = openpyxl.load_workbook(loc)
    sheet=wb_obj.active

    
    for i in range(1,sheet.max_row):
        
        
  
        cli=Expert.query.filter_by(nom=str(sheet["D"][i].value.lower())).first()
    
        if cli is None:
            expert=Expert(genre='M',old=sheet["A"][i].value,nom=sheet["D"][i].value.lower(),numero=sheet["R"][i].value,TYPE=sheet["B"][i].value,
            email=sheet["S"][i].value,email_perso=sheet["T"][i].value,code_tva=sheet["P"][i].value,taux_tva=sheet["Q"][i].value,siret=sheet["K"][i].value,date_entree=sheet["F"][i].value,
            trigramme=sheet["E"][i].value)    
            db.session.add(expert)
            db.session.commit()
            #expert.new="Expert"+str(expert.id)
            #db.session.commit()
            history=Expert_History(expert_id=expert.id,secteur=sheet["C"][i].value,adresse1=sheet["L"][i].value,adresse2=sheet["M"][i].value,cp=sheet["N"][i].value,
            ville=sheet["O"][i].value,login_backof=sheet["U"][i].value,pwd_backof=sheet["V"][i].value,login_extranet=sheet["Y"][i].value,
            pwd_extranet=sheet["Z"][i].value,login_google=sheet["AA"][i].value,pwd_google=sheet["AB"][i].value,observations_de_suivi=sheet["AE"][i].value)
            db.session.add(history)
            db.session.commit()
        else:
            print('already exist')
#talk about date sortie(Format)
        

def tarif_client(loc):
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)

    sheet.cell_value(0, 0)
    for i in range(0,5):
        name=sheet.row_values(i+1)
        tarif =Tarifs.query.filter_by(reference_client=int(name[1])).first()
        cli=Client.query.filter_by(reference=int(name[1])).first()
        r_C=Expert.query.filter_by(nom=str(name[3].lower())).first()
        if r_C is not None:
            rC= r_C.id
        if r_C is None  :
            rC= 0
        r_r=Expert.query.filter_by(nom=str(name[5].lower())).first()
        if r_r is not None:
            rr= r_r.id
        if r_r is None  :
            rr= 0
        dr_a=Expert.query.filter_by(nom=str(name[7].lower())).first()
        if dr_a is not None:
            dra= dr_a.id
        if dr_a is None  :
            dra= 0
        tr_a=Expert.query.filter_by(nom=str(name[9].lower())).first()
        if tr_a is not None:
            tra= tr_a.id
        if tr_a is None  :
            tra= 0
        tr_r=Expert.query.filter_by(nom=str(name[11].lower())).first()
        if tr_r is not None:
            trr= tr_r.id
        if tr_r is None  :
            trr= 0
        tr_s=Expert.query.filter_by(nom=str(name[13].lower())).first()
        if tr_s is not None:
            trs= tr_s.id
        if tr_s is None  :
            trs= 0
        pr_s=Expert.query.filter_by(nom=str(name[15].lower())).first()
        if pr_s is not None:
            prs= pr_s.id
        if pr_s is None  :
            prs= 0
        pr_si=Expert.query.filter_by(nom=str(name[17].lower())).first()
        if pr_si is not None:
            prsi= pr_si.id
        if pr_si is None  :
            prsi= 0
        ra_s=Expert.query.filter_by(nom=str(name[19].lower())).first()
        if ra_s is not None:
            ras= ra_s.id
        if ra_s is None  :
            ras= 0
        if tarif is None:
            if cli :
                taf_base =Tarifs(reference_client=cli.id,code_tva=int(name[2]),
                referent_as_client=rC,com_as_sur_ca_client=name[4],cell_dev_ref_responsable=rr,
                com_cell_dev_ref_responsable=name[6],cell_dev_ref_agent=dra,com_cell_dev_ref_agent=name[8],
                cell_tech_ref_agent=tra,com_cell_tech_Ref_agent=name[10],cell_tech_ref_responsable=trr,
                com_cell_tech_ref_responsable=name[12],cell_tech_ref_suiveur=trs,
                com_cell_tech_ref_suiveur=name[14],cell_planif_ref_responsable=prs,
                com_cell_planif_ref_responsable=name[16],cell_planif_ref_suiveur=prsi,
                com_cell_planif_ref_suiveur=name[18],cell_planif_ref_agent_saisie=ras,
                com_cell_planif_ref_agent_saisie=name[20],taux_meuble=name[21],edl_prix_std=float(name[22]),
                edl_appt_prix_f1=float(name[23]),edl_appt_prix_f2=float(name[24]),edl_appt_prix_f3=float(name[25]),edl_appt_prix_f4=float(name[26]),
                edl_appt_prix_f5=float(name[27]),edl_appt_prix_f6=float(name[28]),edl_pav_villa_prix_t1=float(name[29]), edl_pav_villa_prix_t2=float(name[30]),
                edl_pav_villa_prix_t3=float(name[31]),edl_pav_villa_prix_t4=float(name[32]),edl_pav_villa_prix_t5=float(name[33]),edl_pav_villa_prix_t6=float(name[34]),
                edl_pav_villa_prix_t7=float(name[35]),edl_pav_villa_prix_t8=float(name[36]),chif_appt_prix_stu=float(name[37]),
                chif_appt_prix_f1=float(name[38]),chif_appt_prix_f2=float(name[39]),chif_appt_prix_f3=float(name[40]),
                chif_appt_prix_f4=float(name[41]),chif_appt_prix_f5=float(name[42]),chif_pav_villa_prix_t1=float(name[43]),
                chif_pav_villa_prix_t2=float(name[44]),chif_pav_villa_prix_t3=float(name[45]),chif_pav_villa_prix_t4=float(name[46]),
                chif_pav_villa_prix_t5=float(name[47]),chif_pav_villa_prix_t6=float(name[48]),chif_pav_villa_prix_t7=float(name[49]),
                chif_pav_villa_prix_t8=int(name[50]))
                
                db.session.add(taf_base)
                db.session.commit()
            else:
                print('esit2')
        else:
            print('esit')



























#Missions()






#insert_client('Bailleur')
#insert_client('Locataire')
#insert_client('Prop')
#expert_('Interv')
#expert_('CONCESS')
#expert_('Manager_chiffrage')
#expert_('Agent_chiffrage')
#expert_('Respon Cell Dev')
#expert_('agent Cell Dev')
#expert_('Agent CellTech')
#expert_('Respon Cell Tech')
#expert_('Suiveur Cell Tech')
#expert_('Respon Cell Planif')
#expert_('Suiveur Cell Planif')
#expert_('Agent saisie Cell Planif')
