import sys
import datetime
import openpyxl
import xlrd,xlwt
from Database_project.project.data_base_ import db
from Database_project.project.data_base_.Models import Tarifs,Mission,Client,Expert,Client_History,prospect,prospect_History,Expert_History,Tarif_base
import flask as pd
from Database_project.project.data_base_.client_data  import regex1
from flask import Flask,render_template,url_for,flash,redirect,request,Blueprint,send_from_directory
import os
 
def failed1(av):
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Anomalie')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    #wb.save('C:/Users/user/Downloads/Telegram Desktop/Anomalieexpert.xls')
    filename='Expert_failed.xls'
    file_path = os.path.join(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'),filename)
    loc=str(file_path)
    # set the file path
    #uploaded_file.save(file_path)
    wb.save(loc)
    return send_from_directory(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'), filename,as_attachment=True)

def xperte(loc):
    #wb_obj = openpyxl.load_workbook(loc,data_only=True)
    #anomalie=[]
    wb = xlrd.open_workbook(loc)

    sheet1 = wb.sheet_by_index(0)
    rows=int(sheet1.nrows)
    #sheet=wb_obj.active
    for i in range(1,rows):
        sheet=sheet1.row_values(i)
        if str(sheet[11]) != '' and str(sheet[11]) != 'None':
            cli=Expert.query.filter_by(full=str(sheet[11].lower())).first()
            if cli  is None:
                exp=Expert()
                his=Expert_History()

                exp.nom=regex1(sheet[11],'split')[0]
                exp.full=regex1(sheet[11].lower(),'str3')
                exp.genre=regex1(sheet[10],'str1')
                exp.prenom=regex1(sheet[11],'split')[1]
                exp.TYPE=regex1('Concessionaire','str1')
                db.session.add(exp)
                db.session.commit()
                his.expert_id=exp.id
                db.session.add(his)
                db.session.commit()
        if str(sheet[17]) != '' and str(sheet[17]) != 'None':
            cli=Expert.query.filter_by(full=str(sheet[17].lower())).first()
            if cli  is None:
                exp=Expert()
                his=Expert_History()

                exp.nom=regex1(sheet[17],'split')[0]
                exp.full=regex1(sheet[17].lower(),'str3')
                exp.genre=regex1(sheet[16],'str1')
                exp.prenom=regex1(sheet[17],'split')[1]
                exp.TYPE=regex1('Intervenant','str1')
                db.session.add(exp)
                db.session.commit()
                his.expert_id=exp.id
                db.session.add(his)
                db.session.commit()
        if str(sheet[37]) != '' and str(sheet[37]) != 'None':
            cli=Expert.query.filter_by(full=str(sheet[37].lower())).first()
            if cli  is None:
                exp=Expert()
                his=Expert_History()

                exp.nom=regex1(sheet[37],'split')[0]
                exp.full=regex1(sheet[37].lower(),'str3')
                #exp.genre=regex1(sheet["Q"],'str1')
                exp.prenom=regex1(sheet[37],'split')[1]
                exp.TYPE=regex1('Manager chiffrage','str1')
                db.session.add(exp)
                db.session.commit()
                his.expert_id=exp.id
                db.session.add(his)
                db.session.commit()
        if str(sheet[69]) != '' and str(sheet[69]) != 'None':
            cli=Expert.query.filter_by(full=str(sheet[69].lower())).first()
            if cli  is None:
                exp=Expert()
                his=Expert_History()

                exp.nom=regex1(sheet[69],'split')[0]
                exp.full=regex1(sheet[69].lower(),'str3')
                #exp.genre=regex1(sheet["Q"],'str1')
                exp.prenom=regex1(sheet[69],'split')[1]
                exp.TYPE=regex1('Responsable cell dev','str1')
                db.session.add(exp)
                db.session.commit()
                his.expert_id=exp.id
                db.session.add(his)
                db.session.commit()
        if str(sheet[71]) != '' and str(sheet[71]) != 'None':
            cli=Expert.query.filter_by(full=str(sheet[71].lower())).first()
            if cli  is None:
                exp=Expert()
                his=Expert_History()

                exp.nom=regex1(sheet[71],'split')[0]
                exp.full=regex1(sheet[71].lower(),'str3')
                #exp.genre=regex1(sheet["Q"],'str1')
                exp.prenom=regex1(sheet[71],'split')[1]
                exp.TYPE=regex1('Agent cell dev','str1')
                db.session.add(exp)
                db.session.commit()
                his.expert_id=exp.id
                db.session.add(his)
                db.session.commit()
        if str(sheet[73]) != '' and str(sheet[73]) != 'None':
            cli=Expert.query.filter_by(full=str(sheet[73].lower())).first()
            if cli  is None:
                exp=Expert()
                his=Expert_History()

                exp.nom=regex1(sheet[73],'split')[0]
                exp.full=regex1(sheet[73].lower(),'str3')
                #exp.genre=regex1(sheet["Q"],'str1')
                exp.prenom=regex1(sheet[73],'split')[1]
                exp.TYPE=regex1('Agent cell Tech','str1')
                db.session.add(exp)
                db.session.commit()
                his.expert_id=exp.id
                db.session.add(his)
                db.session.commit()
        if str(sheet[75]) != '' and str(sheet[75]) != 'None':
            cli=Expert.query.filter_by(full=str(sheet[75].lower())).first()
            if cli  is None:
                exp=Expert()
                his=Expert_History()

                exp.nom=regex1(sheet[75],'split')[0]
                exp.full=regex1(sheet[75].lower(),'str3')
                #exp.genre=regex1(sheet["Q"],'str1')
                exp.prenom=regex1(sheet[75],'split')[1]
                exp.TYPE=regex1('Res cell Tech','str1')
                db.session.add(exp)
                db.session.commit()
                his.expert_id=exp.id
                db.session.add(his)
                db.session.commit()

        if str(sheet[77]) != '' and str(sheet[77]) != 'None':
            cli=Expert.query.filter_by(full=str(sheet[77].lower())).first()
            if cli  is None:
                exp=Expert()
                his=Expert_History()

                exp.nom=regex1(sheet[77],'split')[0]
                exp.full=regex1(sheet[77].lower(),'str3')
                #exp.genre=regex1(sheet["Q"],'str1')
                exp.prenom=regex1(sheet[77],'split')[1]
                exp.TYPE=regex1('Suiveur cell Tech','str1')
                db.session.add(exp)
                db.session.commit()
                his.expert_id=exp.id
                db.session.add(his)
                db.session.commit()
        if str(sheet[79]) != '' and str(sheet[79]) != 'None':
            cli=Expert.query.filter_by(full=str(sheet[79].lower())).first()
            if cli  is None:
                exp=Expert()
                his=Expert_History()

                exp.nom=regex1(sheet[79],'split')[0]
                exp.full=regex1(sheet[79].lower(),'str3')
                #exp.genre=regex1(sheet["Q"],'str1')
                exp.prenom=regex1(sheet[79],'split')[1]
                exp.TYPE=regex1('Respon cell Planif','str1')
                db.session.add(exp)
                db.session.commit()
                his.expert_id=exp.id
                db.session.add(his)
                db.session.commit()
        if str(sheet[81]) != '' and str(sheet[81]) != 'None':
            cli=Expert.query.filter_by(full=str(sheet[81].lower())).first()
            if cli  is None:
                exp=Expert()
                his=Expert_History()

                exp.nom=regex1(sheet[81],'split')[0]
                exp.full=regex1(sheet[81].lower(),'str3')
                #exp.genre=regex1(sheet["Q"],'str1')
                exp.prenom=regex1(sheet[81],'split')[1]
                exp.TYPE=regex1('Suiveur cell Planif','str1')
                db.session.add(exp)
                db.session.commit()
                his.expert_id=exp.id
                db.session.add(his)
                db.session.commit()


def xpert(loc):
    try:
        wb_obj = openpyxl.load_workbook(loc,data_only=True)
        anomalie=[]
        sheet=wb_obj.active
        '''if sheet["B"][0].value!='identité agent'and sheet["C"][0].value!='Trigramme'and sheet["D"][0].value!='date entrée'and sheet["P"][0].value!='Ville':
                    return False'''
        for i in range(1,sheet.max_row):
            #n=sheet["M"][i].value
            #if type(n) == str:
                cli=Expert.query.filter_by(full=str(sheet["O"][i].value.lower())).first()
                if cli  is None:
                    exp=Expert()
                    his=Expert_History()
                    
                    exp.nom=regex1(sheet["A"][i].value,'str1')
                    exp.full=regex1(sheet["O"][i].value,'str3')
                    exp.genre=regex1(sheet["P"][i].value,'str1')
                    exp.prenom=regex1(sheet["B"][i].value,'str1')
                    '''numero=regex1(sheet["S"][i].value,'int')
                    if numero == 'Failed':
                        reason="erreur  de numero dans la colonne  Telephone  ,veuillez verifier toute colonne avant d'envoyer"
                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                        sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                        sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                        sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value,reason])
                        continue
                    else:
                        exp.numero=numero'''
                    exp.TYPE=regex1(sheet["D"][i].value,'str1')
                    '''exp.email=regex1(sheet["T"][i].value,'str1')
                    exp.email_perso=regex1(sheet["U"][i].value,'str1')'''
                    exp.code_tva=regex1(sheet["G"][i].value,'str1')
                    taux_tva=regex1(sheet["H"][i].value,'int')   
                    if taux_tva == 'Failed':
                        reason="erreur  de numero dans la colonne  Taux tva ,veuillez verifier toute colonne avant d'envoyer"
                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                        sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                        sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                        sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value,reason])
                        continue 
                    else:
                        exp.taux_tva=taux_tva
                    siret=regex1(sheet["F"][i].value,'int')
                    if siret == 'Failed':
                        reason="erreur  de numero dans la colonne  Siret ,veuillez verifier toute colonne avant d'envoyer"
                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                        sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                        sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                        sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value,reason])
                        continue 
                    else:
                        exp.siret=siret
                    date_entree=regex1(sheet["E"][i].value,'date')
                    if date_entree == False:
                        reason="erreur  de numero dans la colonne  date entree ,veuillez verifier toute colonne avant d'envoyer"
                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                        sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                        sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                        sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value,reason])
                        continue 
                    else:
                        exp.date_entree=date_entree
                    exp.trigramme=regex1(sheet["C"][i].value,'str1')    
                    
                    his.secteur=regex1(sheet["J"][i].value,'str1') 
                    his.adresse1=regex1(sheet["K"][i].value,'str1') 
                    his.adresse2=regex1(sheet["L"][i].value,'str1')
                    cp=regex1(sheet["M"][i].value,'int') 
                    if cp=='Failed':
                        reason="erreur  de numero dans la colonne  CP ,veuillez verifier toute colonne avant d'envoyer"
                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                        sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                        sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                        sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value,reason])
                        continue 
                    else:
                        his.cp=cp 
                    his.ville=regex1(sheet["N"][i].value,'str1') 
                    '''his.login_backof=regex1(sheet["V"][i].value,'str1') 
                    his.pwd_backof=regex1(sheet["W"][i].value,'str1') 
                    his.login_extranet=regex1(sheet["Z"][i].value,'str1') 
                    his.pwd_extranet=regex1(sheet["AA"][i].value,'str1') 
                    his.login_google=regex1(sheet["X"][i].value,'str1') 
                    his.pwd_google=regex1(sheet["AB"][i].value,'str1') 
                    his.observations_de_suivi=regex1(sheet["AD"][i].value,'str1')'''
                    his.actif_parti=regex1(sheet["I"][i].value,'str1') 
                    '''his.type_certification=regex1(sheet["H"][i].value,'str1') 
                    date_certification_initiale=regex1(sheet["F"][i].value,'date') #CHECK FOR IF STATE
                    if date_certification_initiale == False:
                        reason="erreur  de numero dans la colonne  date_certification_initiale  ,veuillez verifier toute colonne avant d'envoyer"
                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                        sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                        sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                        sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value,reason])
                        continue 
                    else:
                        his.date_certification_initiale=date_certification_initiale
                    date_renouv_certification=regex1(sheet["G"][i].value,'date') #CHECK FOR IF STATE
                    if date_renouv_certification == False:
                        reason="erreur  de numero dans la colonne  date_renouv_certification  ,veuillez verifier toute colonne avant d'envoyer"
                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                        sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                        sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                        sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value,reason])
                        continue 
                    else:
                        his.date_renouv_certification=date_renouv_certification'''
                    ''''his.pwd_gsuite=regex1(sheet["AB"][i].value,'date') '''
                    db.session.add(exp)
                    db.session.add(his)
                    db.session.commit()
                    his.expert_id=exp.id
                    db.session.commit()
                else:
                    print('existe')
                
        if anomalie == []:
            return True
        else:
            return failed1(anomalie)
    except:
        return 'Fake'
