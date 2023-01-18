import sys
import datetime
import openpyxl
import xlrd,xlwt
from Database_project.project.data_base_ import db,create_app
from Database_project.project.data_base_.Models import Tarifs,Mission,Client,Expert,Client_History,prospect,prospect_History,Expert_History,Tarif_base
import flask as pd
from flask import Flask,render_template,url_for,flash,redirect,request,Blueprint,make_response,send_from_directory
import os



def regex1(data,Type):
    if  Type=='str':
        if type(data)==str:
            if data.isalnum() == False:
                a=str(data).split()
                vii=''.join(a)
                vii=vii.split("-")
                vii=''.join(vii)
                return data
            if data == 'None':
                return ''
            if data == 'xxx':
                return ''
            if data == 'XXX':
                return ''
            if data == None:
                return ''
            else:
                return data
        else:
            return ''
    
    if  Type=='str1':
        if type(data)==str:
            if data == 'xxx':
                return ''
            if data == 'XXX':
                return ''
            if data == 'None':
                return ''
            if data == None:
                return ''
            else:
                return data
        else:
            return ''
    if  Type=='str3':
        if type(data)==str:
            if data == 'xxx':
                return ''
            if data == 'XXX':
                return ''
            if data == 'None':
                return ''
            if data == None:
                return ''
            else:
                return data.lower()
                
        else:
            return ''
    if  Type=='split':
        if type(data)==str and data != '':
            nom=data.split()
            if len(nom)>=3:
                nom=[nom[0],nom[1]+''+nom[2]]
                return nom
            if len(nom)>=2:
                return nom
            else:
                last=[nom[0],'']
                return last

        else:
            return ['','']

    if  Type=='int':
        if type(data) == int:
            return int(data)
        if type(data) == str:
            if data.isnumeric() == True:
                return int(data)
            a="".join(data.split())
            if a.isnumeric() == True:
                return int(a) 
            else:
                return 'Failed'
        if data == None:
            return 0 
        try:

            return int(data)
        except:
            return 'Failed'
        
    if  Type=='M':
        if data == '':
            return None
        if type(data) == int:
            return int(data)
        if type(data) == float:
            return int(data)
        if type(data) == str:
            if data.isnumeric() == True:
                return int(data)
            a="".join(data.split())
            if a.isnumeric() == True:
                return int(a) 
            else:
                return 0
        if data == None:
            return None 
        else:
            return 0            
            
    
    if  Type=='dec':
        
        if isinstance(data,int) == True:
            return round(data,2)
        if isinstance(data,float) == True:
            return data
        if data == '':
            return 0.00
        if data == None:
            return 0.00
        else:
            return 'Failed'
    
    if  Type=='S':
        if isinstance(data,int) == True:
            return round(data,2)
        if isinstance(data,float) == True:
            return data
        if data == '':
            return 0.00
        if type(data)==str:
            return 0.00
        if data == None:
            return 0.00
        else:
            return 0.00

    if  Type == 'perc':
        if type(data) == str:
            if data == '':
                return 0.00
            if data[-1] == '%':
                return float(data.strip('%'))/100
            else:
                return 0.00
        if type(data) == int:
            return data/100
        if type(data) == float:
            if data % 1 !=0:
                return data
            else:
                return data/100
        else:
                return 0.00
    if  Type=='date':
        
        if isinstance(data,datetime.datetime) == True:
            return data
        elif isinstance(data,datetime.date) == True:
            return data
        elif data == None:
            a=1
        else:
            
            return False
            

def failed(av):
    
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Anomalie')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    filename='client%prospect_failed.xls'
    file_path = os.path.join(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'),filename)
    loc=str(file_path)
    # set the file path
    #uploaded_file.save(file_path)
    wb.save(loc)
    return send_from_directory(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'), filename,as_attachment=True)
    #wb.save('C:/Users/user/Downloads/Telegram Desktop/Anomalieclient.xls')

def failed1(av):
    
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Anomalie')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    filename='prospect_failed.xls'
    file_path = os.path.join(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'),filename)
    loc=str(file_path)
    # set the file path
    #uploaded_file.save(file_path)
    wb.save(loc)
    return send_from_directory(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'), filename,as_attachment=True) 
    #wb.save('C:/Users/user/Downloads/Telegram Desktop/AnomaliePROSPECT.xls')

def good1(av):
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Bien')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    filename='prospect_good.xls'
    file_path = os.path.join(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'),filename)
    loc=str(file_path)
    # set the file path
    #uploaded_file.save(file_path)
    wb.save(loc)
    return send_from_directory(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'), filename,as_attachment=True)
    #wb.save('C:/Users/user/Downloads/Telegram Desktop/BienPROSPECT.xls')

def good2(av):
    ba=[]
    v=0
    for oo in av:
        tr=len(oo)
    for i in range(0,tr):
        ba.append(i)
    wb = xlwt.Workbook()
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    ws = wb.add_sheet('Bien')
    for oo in av:
            for q,i in zip(ba,oo) :
                if isinstance(i,datetime.datetime) == True:
                    ws.write(v, q, i,style1)
                elif isinstance(i,datetime.date) == True:
                    ws.write(v, q, i,style1) 
                else:
                    ws.write(v, q, i)
            v=v+1
    filename='client_good.xls'
    file_path = os.path.join(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'),filename)
    loc=str(file_path)
    # set the file path
    #uploaded_file.save(file_path)
    wb.save(loc)
    return send_from_directory(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static','export'), filename,as_attachment=True)
    #wb.save('C:/Users/user/Downloads/Telegram Desktop/BienClient.xls')


def date_(floa,date):
        try:
            v=datetime.datetime(*xlrd.xldate_as_tuple(floa,date))
            return v
        except:
            return floa

def lienta(loc):
    wb = xlrd.open_workbook(loc)

    sheet = wb.sheet_by_index(0)
    rows=int(sheet.nrows)
    #wb_obj = openpyxl.load_workbook(loc,data_only=True)
    anomalie=[]
   # sheet=wb_obj.active
    for i in range(0,rows):
        sheet1=sheet.row_values(i)
        v=sheet1[0]
        try:
            int(v) 
            
            cli=Client.query.filter_by(reference=int(sheet1[0])).first()
            if cli is None:
                create=Client()
                
                Histo=Client_History()
                
                create.reference=int(sheet1[0])
                create.TYPE=regex1(sheet1[12],'str')
                create.societe=regex1(sheet1[1],'str')
                #create.enseigne=regex1(sheet1["C"],'str')
                create.titre=regex1(sheet1[2],'str')
                create.nom=regex1(sheet1[3],'split')[0]
                create.prenom=regex1(sheet1[3],'split')[1]
                #numero=regex1(sheet1[8],'int')
                #if numero == 'Failed':
                    #print('failed')
                    #'''reason='erreur  de numero dans la colonne  Tel principal client ,veuillez verifier toute colonne avant d"envoyer'
                    #anomalie.append([sheet1["A"],sheet1["B"],sheet1["C"],
                    #sheet1["D"],sheet1["E"],sheet1["F"],sheet1["G"],
                    #sheet1["H"],sheet1["I"],sheet1["J"],sheet1["K"],
                    #sheet1["L"],sheet1["M"],sheet1["N"],sheet1["O"],
                    #sheet1["P"],sheet1["Q"],reason])'''
                    #continue
                #else:
                create.numero=0
                db.session.add(create)
                db.session.commit()
                cp=regex1(sheet1[6],'int')
                if cp == 'Failed':
                   print('failed')
                   ''' reason='erreur  de numero dans la colonne  code postal ,veuillez verifier toute colonne avant d"envoyer'
                    anomalie.append([sheet1["A"],sheet1["B"],sheet1["C"],
                    sheet1["D"],sheet1["E"],sheet1["F"],sheet1["G"],
                    sheet1["H"],sheet1["I"],sheet1["J"],sheet1["K"],
                    sheet1["L"],sheet1["M"],sheet1["N"],sheet1["O"],
                    sheet1["P"],sheet1["Q"],reason])
                    continue  ''' 
                else:
                    Histo.cp=cp
                Histo.adresse1=regex1(sheet1[4],'str1')
                Histo.ville=regex1(sheet1[7],'str1')
                Histo.client_id=create.id
                db.session.add(Histo)
                db.session.commit()
        except:
            print('ok')
        
    if anomalie == []:
            return True
    else:
        return failed(anomalie)

def lient(loc):
    #try:
        anomalie=[]
        anomalie1=[]
        Bien1=[]
        Bien=[]
        a=[]
        b=[]
        p=0
        wb_obj = openpyxl.load_workbook(loc,data_only=True)
        sheet=wb_obj.active
        '''if sheet["W"][0].value!='CODE POSTAL'and sheet["AB"][0].value!='Tel principal client'and sheet["J"][0].value!='Ref code client- service comptabilité'and sheet["P"][0].value!='Fonction responsable':
                    return False'''
        for i in range(0,sheet.max_row):
                if sheet["M"][i].value!='PROSPECT':
                    v=sheet["A"][i].value
                    
                    if type(v)==int: 
                        if v  not in a:
                            cli=Client.query.filter_by(reference=int(sheet["A"][i].value)).first()
                            #a.append(v)
                            if cli is None:
                                create=Client()
                                
                                Histo=Client_History()
                                
                                create.reference=int(sheet["A"][i].value)
                                create.TYPE=regex1(sheet["M"][i].value,'str')
                                create.societe=regex1(sheet["B"][i].value,'str')
                                create.enseigne=regex1(sheet["C"][i].value,'str')
                                create.titre=regex1(sheet["D"][i].value,'str')
                                create.nom=regex1(sheet["E"][i].value,'str1')
                                create.prenom=regex1(sheet["F"][i].value,'str1')
                                numero=regex1(sheet["H"][i].value,'int')
                                if numero == 'Failed':
                                    reason='erreur  de numero dans la colonne  Tel principal client ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,reason])
                                    continue
                                else:
                                    create.numero=numero
                                
                                siret=regex1(sheet["I"][i].value,'int')
                                if siret == 'Failed':
                                    reason='erreur  de numero dans la colonne  Siret ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,reason])
                                    continue
                                else:
                                    create.siret=siret
                                create.email=regex1(sheet["G"][i].value,'str1')
                                date_creation=regex1(sheet["J"][i].value,'date')
                                if date_creation == False:
                                    reason='erreur  de numero dans la colonne  date creation ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,reason])
                                    continue
                                else:
                                    create.date_creation=date_creation
                                
                                Histo.adresse1=regex1(sheet["K"][i].value,'str1')
                                '''Histo.adresse2=regex1(sheet["V"][i].value,'str1')'''
                                Histo.etat_client=regex1(sheet["L"][i].value,'str')
                                cp=regex1(sheet["N"][i].value,'int')
                                if cp == 'Failed':
                                    reason='erreur  de numero dans la colonne  code postal ,veuillez verifier toute colonne avant d"envoyer'
                                    anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,reason])
                                    continue   
                                else:
                                    Histo.cp=cp
                                Histo.ville=regex1(sheet["O"][i].value,'str1')
                                Histo.pays='France'
                                Histo.login_extranet=regex1(sheet["P"][i].value,'str1')
                                Histo.mpd_extranet=regex1(sheet["Q"][i].value,'str1')
                                db.session.add(create)
                                db.session.add(Histo)
                                db.session.commit()
                                Histo.client_id=create.id
                                db.session.commit()
                                '''Bien.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value,sheet["R"][i].value,sheet["S"][i].value,
                                    sheet["T"][i].value,sheet["U"][i].value,sheet["V"][i].value,sheet["W"][i].value,
                                    sheet["X"][i].value,sheet["Y"][i].value,sheet["Z"][i].value,sheet["AA"][i].value
                                    ,sheet["AB"][i].value,sheet["AC"][i].value])'''
                            else:
                                reason="erreur  de doublon dans la colonne  reference,veuillez verifier toute colonne avant d'envoyer"
                                anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                sheet["P"][i].value,sheet["Q"][i].value,reason])
                        else:
                            reason="erreur  de doublon dans la colonne  reference,veuillez verifier toute colonne avant d'envoyer"
                            anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                            sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                            sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                            sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                            sheet["P"][i].value,sheet["Q"][i].value,reason])
                    else:
                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                        sheet["P"][i].value,sheet["Q"][i].value,"ERREUR REFERENCE"])
                if  sheet["M"][i].value =='PROSPECT':
                        v=sheet["A"][i].value 
                        if type(v)==int: 
                            if v  not in b:
                                cli=prospect.query.filter_by(reference=int(sheet["A"][i].value)).first()
                                b.append(v)
                                if cli is None:
                                    create=prospect()
                                    
                                    Histo=prospect_History()
                                    
                                    create.reference=int(sheet["A"][i].value)
                                    create.TYPE=regex1(sheet["M"][i].value,'str')
                                    create.societe=regex1(sheet["B"][i].value,'str')
                                    create.enseigne=regex1(sheet["C"][i].value,'str')
                                    create.titre=regex1(sheet["D"][i].value,'str')
                                    create.nom=regex1(sheet["E"][i].value,'str1')
                                    create.prenom=regex1(sheet["F"][i].value,'str1')
                                    numero=regex1(sheet["H"][i].value,'int')
                                    if numero == 'Failed':
                                        reason='erreur  de numero dans la colonne  Tel principal client ,veuillez verifier toute colonne avant d"envoyer'
                                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                        sheet["P"][i].value,sheet["Q"][i].value,reason])
                                        continue
                                    else:
                                        create.numero=numero
                                    
                                    siret=regex1(sheet["I"][i].value,'int')
                                    if siret == 'Failed':
                                        reason='erreur  de numero dans la colonne  Siret ,veuillez verifier toute colonne avant d"envoyer'
                                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                        sheet["P"][i].value,sheet["Q"][i].value,reason])
                                        continue
                                    else:
                                        create.siret=siret
                                    create.email=regex1(sheet["G"][i].value,'str1')
                                    date_creation=regex1(sheet["J"][i].value,'date')
                                    if date_creation == False:
                                        reason='erreur  de numero dans la colonne  date creation ,veuillez verifier toute colonne avant d"envoyer'
                                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                        sheet["P"][i].value,sheet["Q"][i].value,reason])
                                        continue
                                    else:
                                        create.date_creation=date_creation
                                    
                                    Histo.adresse1=regex1(sheet["K"][i].value,'str1')
                                    '''Histo.adresse2=regex1(sheet["V"][i].value,'str1')'''
                                    Histo.etat_client=regex1(sheet["L"][i].value,'str')
                                    cp=regex1(sheet["N"][i].value,'int')
                                    if cp == 'Failed':
                                        reason='erreur  de numero dans la colonne  code postal ,veuillez verifier toute colonne avant d"envoyer'
                                        anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                        sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                        sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                        sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                        sheet["P"][i].value,sheet["Q"][i].value,reason])
                                        continue   
                                    else:
                                        Histo.cp=cp
                                    Histo.ville=regex1(sheet["O"][i].value,'str1')
                                    Histo.pays='France'
                                    Histo.login_extranet=regex1(sheet["P"][i].value,'str1')
                                    Histo.mpd_extranet=regex1(sheet["Q"][i].value,'str1')
                                    db.session.add(create)
                                    db.session.add(Histo)
                                    db.session.commit()
                                    Histo.prospect=create.id
                                    db.session.commit()
                                    '''Bien1.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                    sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                    sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                    sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                    sheet["P"][i].value,sheet["Q"][i].value])'''
                                else:
                                    ("Anomalie doublon")
                            else:
                                reason="Anomalie doublon sur la reference"
                                anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                                sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                                sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                                sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                                sheet["P"][i].value,sheet["Q"][i].value,reason])
                        else:
                            anomalie.append([sheet["A"][i].value,sheet["B"][i].value,sheet["C"][i].value,
                            sheet["D"][i].value,sheet["E"][i].value,sheet["F"][i].value,sheet["G"][i].value,
                            sheet["H"][i].value,sheet["I"][i].value,sheet["J"][i].value,sheet["K"][i].value,
                            sheet["L"][i].value,sheet["M"][i].value,sheet["N"][i].value,sheet["O"][i].value,
                            sheet["P"][i].value,sheet["Q"][i].value,"erreur dans la colonne  reference,veuillez verifier toute colonne avant d'envoyer"])
        if anomalie == []:
            return True
        else:
            return failed(anomalie)
        '''if anomalie1 == []:
            print(anomalie1)
        else:
            return failed1(anomalie1)
        if Bien == []:
            print(Bien)
        else:
            return good2(Bien)
        if Bien1 == []:
            print(Bien1)
        else:
            return good1(Bien1)'''
    #except:
   #     return 'Fake'



